#Who is topper:-
from openpyxl import load_workbook
from prettytable import PrettyTable
from colorama import init,Fore,Style
init(autoreset=True) #automatically resets colors after each print
from sys import argv


#importing a excel data
 
if len(argv) < 2:
    print("Usage Python Script...")
    exit()
    

input_file = argv[1]




class Node:
    def __init__(self,s_no, roll_no, name, rank):
        
        self.s_no = s_no
        self.roll_no = roll_no  
        self.name = name
        self.rank = rank
        self.next = None
         
        
        
class stack:
    
    def __init__(self):
        self.top = None
        self.count = 0
         
        
        
    def is_empty(self):
        return self.top == None
        
        
    def push(self,*args):
        if isinstance(args[0],Node):
            new_node = args[0]
        else:
            
            new_node = Node(*args)
        #link to new_node to exist node
        new_node.next = self.top
        self.top = new_node
        self.count+=1
                
            
    def pop(self):
        #delete a node in a stack
        if self.is_empty():
            return None
        
        else:
            poped_node = self.top
            self.top = self.top.next
            self.count -= 1
            return poped_node
            
            
    def peek(self):
        #using ternary style
        return self.top.rank if self.top else None
            
            
    
    def get_data(self):
        try:
            #opening a excel file name is {leaderboard.xlsx}
            workbook = load_workbook("leaderboard.xlsx")
            sheet = workbook.active
        
        #if file not found in directory
        except FileNotFoundError:
            print(f"ERROR file not found {input_file}")
            exit()
            
                
        for row in sheet.iter_rows(min_row=2 , values_only=True):
            #adding a row in these 4 different variables
            s_no,roll_no,name,rank = row
            self.push(s_no,roll_no,name,int(rank))   
        
        self.sort()
            
            
    def sort(self):
        #making a temp stack to sorted a list based on Given percentage
        temp = stack()
        while not self.is_empty():
            
            curr = self.pop() #curr is a node object
            
            while not temp.is_empty() and temp.peek() < curr.rank:
                
                self.push(temp.pop())
                
            temp.push(curr) #push the node object directly
            
        while not temp.is_empty():
            self.push(temp.pop())
            
            
            
    def traverse(self):
        if self.top == None:
            print( "Stack is Empty")
            return 
        
        else:
            curr = self.top
            count = 1
            
            #Headings Of Table
            table = PrettyTable(["Sr.No.","Roll-No","Name","Rank"])
            
            
            
            #color combinations:-
            GREEN = Fore.GREEN
            YEELOW = Fore.YELLOW
            BLUE = Fore.BLUE
            RESET = Style.RESET_ALL
            
        while curr:
            
            #for first position just like me{GREEN}
            #different Position differ color
            
            if count == 1:
                table.add_row([f"{GREEN}{count}{RESET}",
                               f"{GREEN}{curr.roll_no}{RESET}",
                               f"{GREEN}{curr.name}{RESET}",
                               f"{GREEN}{curr.rank}{RESET}"])
            
            elif count == 2:
                table.add_row([f"{YEELOW}{count}{RESET}",
                               f"{YEELOW}{curr.roll_no}{RESET}",
                               f"{YEELOW}{curr.name}{RESET}",
                               f"{YEELOW}{curr.rank}{RESET}"])
            
            elif count == 3:
                table.add_row([f"{BLUE}{count}{RESET}",
                               f"{BLUE}{curr.roll_no}{RESET}",
                               f"{BLUE}{curr.name}{RESET}",
                               f"{BLUE}{curr.rank}{RESET}"])
                
            else:
                table.add_row([count, curr.roll_no, curr.name, curr.rank])
            
    
            count+=1    
            curr = curr.next
       
        print(table)
            
            
            
             
                
s = stack()
s.get_data()
s.traverse()



#Who is topper is a Python script Read Details Git-hub readme file

#Using First Time colorama use in Python script to Visible color in CMD/Powershell.
#Using First Time Open Pyxl module To deal with get a data in excel file.
#Using Python Table format Module <prettytable>.

#It's small part of data analysis field but Next i give better and honestly i say some thing i ask form chatgpt because it's new for me {Sorry}


                

            
            
            
     
